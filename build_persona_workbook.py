#!/usr/bin/env python3
"""Build a persona-grouped workbook from the classified rows."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = json.load(open("/tmp/persona_rows.json"))  # (fn, ln, title, company, persona)

FONT = "Arial"
TARGET = [
    "Marketing / Digital / Brand",
    "Martech / Marketing Ops & Data",
    "CX / Customer Experience & AI",
]
OTHER = [
    "Product & Design",
    "Operations & Delivery",
    "Risk / Compliance",
    "Commercial / Sales / Eng / Strategy / Other",
]
COLORS = {
    "Marketing / Digital / Brand": "DDEBF7",
    "Martech / Marketing Ops & Data": "E2EFDA",
    "CX / Customer Experience & AI": "FCE4D6",
    "Product & Design": "FFF2CC",
    "Operations & Delivery": "EDEDED",
    "Risk / Compliance": "F2DCDB",
    "Commercial / Sales / Eng / Strategy / Other": "F4F4F4",
}

from collections import Counter, defaultdict
counts = Counter(r[4] for r in rows)
by_p = defaultdict(list)
for fn, ln, title, company, p in rows:
    by_p[p].append((company, fn, ln, title))
for p in by_p:
    by_p[p].sort(key=lambda x: (x[0], x[2]))

hdr_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
cell_font = Font(name=FONT, size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


wb = Workbook()

# ---- Summary sheet ----
ws = wb.active
ws.title = "Summary"
total = len(rows)
ws["A1"] = "Prospect Personas - London Finance Conference"
ws["A1"].font = Font(name=FONT, bold=True, size=14)
ws["A3"] = f"Total contacts: {total}   |   Companies: {len(set(r[3] for r in rows))}"
ws["A3"].font = Font(name=FONT, size=10)

ws["A5"] = "THREE MARKETING-RELEVANT PERSONAS (your matrix) -> one message template each:"
ws["A5"].font = Font(name=FONT, bold=True, size=11)
r = 6
tgt_total = 0
for p in TARGET:
    ws.cell(r, 1, p).font = Font(name=FONT, bold=True, size=10)
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=COLORS[p])
    ws.cell(r, 2, counts[p]).font = cell_font
    ws.cell(r, 3, f"{100*counts[p]/total:.0f}%").font = cell_font
    tgt_total += counts[p]
    r += 1
ws.cell(r, 1, "Subtotal (3 target personas)").font = Font(name=FONT, bold=True, size=10)
ws.cell(r, 2, tgt_total).font = Font(name=FONT, bold=True, size=10)
ws.cell(r, 3, f"{100*tgt_total/total:.0f}%").font = Font(name=FONT, bold=True, size=10)
r += 2

ws.cell(r, 1, "OTHER PERSONAS (outside the 3 marketing templates - may need a 4th track or deprioritising):").font = Font(name=FONT, bold=True, size=11)
r += 1
for p in OTHER:
    ws.cell(r, 1, p).font = cell_font
    ws.cell(r, 1).fill = PatternFill("solid", fgColor=COLORS[p])
    ws.cell(r, 2, counts[p]).font = cell_font
    ws.cell(r, 3, f"{100*counts[p]/total:.0f}%").font = cell_font
    r += 1
r += 1
note = [
    "Most predominant OVERALL: 1) Marketing/Digital, 2) Product & Design, 3) Operations.",
    "Most predominant MARKETING personas (your matrix): Marketing/Digital > CX/AI > Martech.",
    "Persona assigned from job title (keyword rules + manual fixes) - first pass, easy to adjust.",
    "Each target persona has its own tab below; 'Other personas' tab holds the rest for reference.",
]
for line in note:
    ws.cell(r, 1, line).font = Font(name=FONT, italic=True, size=9)
    r += 1
ws.column_dimensions["A"].width = 92
ws.column_dimensions["B"].width = 8
ws.column_dimensions["C"].width = 8

# ---- One tab per target persona ----
def add_sheet(name, personas):
    ws = wb.create_sheet(name)
    style_header(ws, ["Persona", "Company", "First Name", "Last Name", "Job Title"])
    rr = 2
    for p in personas:
        for company, fn, ln, title in by_p.get(p, []):
            vals = [p, company, fn, ln, title]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(rr, c, v)
                cell.font = cell_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")
            ws.cell(rr, 1).fill = PatternFill("solid", fgColor=COLORS[p])
            rr += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{rr-1}"
    for col, w in {"A": 30, "B": 20, "C": 16, "D": 20, "E": 66}.items():
        ws.column_dimensions[col].width = w
    return rr - 2

add_sheet("1. Marketing-Digital", ["Marketing / Digital / Brand"])
add_sheet("2. Martech", ["Martech / Marketing Ops & Data"])
add_sheet("3. CX-AI", ["CX / Customer Experience & AI"])
add_sheet("Other personas", OTHER)

out = "/home/user/ddalgin.github.io/London_Finance_Personas.xlsx"
wb.save(out)
print("Saved", out)
for p in TARGET + OTHER:
    print(f"  {counts[p]:3d}  {p}")
