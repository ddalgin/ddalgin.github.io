#!/usr/bin/env python3
"""Flat prospect sheet (all 309) with a Segment column: 3 personas + catch-all."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from build_prospect_list import DATA as BASE_DATA
from country_overrides import OVERRIDES

rows = json.load(open("/tmp/persona_rows.json"))  # (fn, ln, title, company, persona)

# email + country lookups (only exist for the original 220)
email_lookup, country_lookup = {}, {}
def split_name(full):
    parts = full.split()
    return (parts[0], " ".join(parts[1:])) if len(parts) > 1 else (full, "")
for company, people in BASE_DATA:
    for full, title, mail in people:
        email_lookup[(company, full)] = mail
for k, (ctry, conf, ev) in OVERRIDES.items():
    country_lookup[k] = (ctry, conf)

# persona -> segment (3 target + catch-all)
SEG = {
    "Marketing / Digital / Brand": "1 - Marketing / Digital / Brand",
    "Martech / Marketing Ops & Data": "2 - Martech / Marketing Ops & Data",
    "CX / Customer Experience & AI": "3 - CX / Customer Experience & AI",
}
CATCHALL = "4 - Other (non-marketing)"
SEG_COLOR = {
    "1 - Marketing / Digital / Brand": "DDEBF7",
    "2 - Martech / Marketing Ops & Data": "E2EFDA",
    "3 - CX / Customer Experience & AI": "FCE4D6",
    "4 - Other (non-marketing)": "EDEDED",
}

data = []
for fn, ln, title, company, persona in rows:
    seg = SEG.get(persona, CATCHALL)
    full = f"{fn} {ln}".strip()
    email = email_lookup.get((company, full), "")
    ctry, conf = country_lookup.get(f"{company}||{full}", ("", ""))
    data.append([fn, ln, title, company, seg, email, ctry, conf])

# Keep the exact order of the user's external Excel. persona_rows.json is already
# in that order (companies in list order), except Jennifer Scott was appended with
# the new contacts - in the external sheet she sits inside NatWest after Tanvi Gokhali.
def _find(pred):
    for i, d in enumerate(data):
        if pred(d):
            return i
    return None
_ji = _find(lambda d: d[0] == "Jennifer" and d[3] == "NatWest Group")
_jen = data.pop(_ji)
_ti = _find(lambda d: d[0] == "Tanvi" and d[3] == "NatWest Group")
data.insert(_ti + 1, _jen)

FONT = "Arial"
HEADERS = ["First Name", "Last Name", "Job Title", "Company", "Segment",
           "Email", "Country", "Country Confidence"]
wb = Workbook()
ws = wb.active
ws.title = "Prospects (segmented)"
hf = Font(name=FONT, bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1F4E78")
cf = Font(name=FONT, size=10)
thin = Side(style="thin", color="D9D9D9")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(1, c, h)
    cell.font = hf; cell.fill = hfill; cell.border = bd
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r = 2
for row in data:
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.font = cf; cell.border = bd
        cell.alignment = Alignment(vertical="center")
    ws.cell(r, 5).fill = PatternFill("solid", fgColor=SEG_COLOR[row[4]])  # Segment
    r += 1
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{r-1}"
for col, w in {"A":16,"B":20,"C":58,"D":18,"E":34,"F":38,"G":16,"H":16}.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 28

# --- Segment summary sheet ---
from collections import Counter
sc = Counter(row[4] for row in data)
s = wb.create_sheet("Segment counts")
s["A1"] = "Segment counts"; s["A1"].font = Font(name=FONT, bold=True, size=13)
rr = 3
tot = len(data)
for seg in ["1 - Marketing / Digital / Brand","2 - Martech / Marketing Ops & Data",
            "3 - CX / Customer Experience & AI","4 - Other (non-marketing)"]:
    s.cell(rr,1,seg).font = Font(name=FONT, bold=True, size=10)
    s.cell(rr,1).fill = PatternFill("solid", fgColor=SEG_COLOR[seg])
    s.cell(rr,2,sc[seg]).font = cf
    s.cell(rr,3,f"{100*sc[seg]/tot:.0f}%").font = cf
    rr += 1
s.cell(rr,1,"TOTAL").font = Font(name=FONT, bold=True, size=10)
s.cell(rr,2,tot).font = Font(name=FONT, bold=True, size=10)
s.column_dimensions["A"].width = 40
s.column_dimensions["B"].width = 8
s.column_dimensions["C"].width = 8

out = "/home/user/ddalgin.github.io/London_Finance_Prospects_Segmented.xlsx"
wb.save(out)
print("Saved", out, "rows:", len(data))
for seg in ["1 - Marketing / Digital / Brand","2 - Martech / Marketing Ops & Data",
            "3 - CX / Customer Experience & AI","4 - Other (non-marketing)"]:
    print(f"  {sc[seg]:3d}  {seg}")
print("  with email:", sum(1 for d in data if d[5]), "/ with country:", sum(1 for d in data if d[6]))
