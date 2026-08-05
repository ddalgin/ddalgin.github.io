#!/usr/bin/env python3
"""Flat prospect sheet (all 309) in the external-Excel row order, with a 6-department
Segment column (+ residual 'Other')."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from build_prospect_list import DATA as BASE_DATA
from country_overrides import OVERRIDES

seg_rows = json.load(open("/tmp/segment6_rows.json"))  # (fn, ln, title, company, segment)

# email + country lookups (exist only for the original 220)
email_lookup = {}
for company, people in BASE_DATA:
    for full, title, mail in people:
        email_lookup[(company, full)] = mail

SEG_COLOR = {
    "Marketing": "DDEBF7",
    "CX / Operations": "FCE4D6",
    "Product": "FFF2CC",
    "Digital & Martech": "E2EFDA",
    "Compliance": "F2DCDB",
    "AI": "E4DFEC",
    "Other (outside these 6)": "EDEDED",
}
ORDER = ["Marketing", "CX / Operations", "Product", "Digital & Martech",
         "Compliance", "AI", "Other (outside these 6)"]

data = []
for fn, ln, title, company, seg in seg_rows:
    full = f"{fn} {ln}".strip()
    email = email_lookup.get((company, full), "")
    ov = OVERRIDES.get(f"{company}||{full}")
    ctry, conf = (ov[0], ov[1]) if ov else ("", "")
    data.append([fn, ln, title, company, seg, email, ctry, conf])

# Keep external-Excel order: seg_rows is already in that order except Jennifer Scott,
# who sits inside NatWest (after Tanvi Gokhali) in the external sheet.
def _find(pred):
    for i, d in enumerate(data):
        if pred(d):
            return i
_ji = _find(lambda d: d[0] == "Jennifer" and d[3] == "NatWest Group")
_jen = data.pop(_ji)
_ti = _find(lambda d: d[0] == "Tanvi" and d[3] == "NatWest Group")
data.insert(_ti + 1, _jen)

FONT = "Arial"
HEADERS = ["First Name", "Last Name", "Job Title", "Company", "Segment",
           "Email", "Country", "Country Confidence"]
wb = Workbook()
ws = wb.active
ws.title = "Prospects (6 segments)"
hf = Font(name=FONT, bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1F4E78")
cf = Font(name=FONT, size=10)
thin = Side(style="thin", color="D9D9D9")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
for c, h in enumerate(HEADERS, 1):
    cell = ws.cell(1, c, h)
    cell.font = hf; cell.fill = hfill; cell.border = bd
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r = 2
for row in data:
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v)
        cell.font = cf; cell.border = bd
        cell.alignment = Alignment(vertical="center")
    ws.cell(r, 5).fill = PatternFill("solid", fgColor=SEG_COLOR[row[4]])
    r += 1
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:H{r-1}"
for col, w in {"A":16,"B":20,"C":58,"D":18,"E":24,"F":38,"G":16,"H":16}.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 28

# summary sheet
from collections import Counter
sc = Counter(row[4] for row in data)
s = wb.create_sheet("Segment counts")
s["A1"] = "Six-department segmentation"; s["A1"].font = Font(name=FONT, bold=True, size=13)
rr = 3
tot = len(data)
for seg in ORDER:
    s.cell(rr, 1, seg).font = Font(name=FONT, bold=True, size=10)
    s.cell(rr, 1).fill = PatternFill("solid", fgColor=SEG_COLOR[seg])
    s.cell(rr, 2, sc[seg]).font = cf
    s.cell(rr, 3, f"{100*sc[seg]/tot:.0f}%").font = cf
    rr += 1
s.cell(rr, 1, "TOTAL").font = Font(name=FONT, bold=True, size=10)
s.cell(rr, 2, tot).font = Font(name=FONT, bold=True, size=10)
s.column_dimensions["A"].width = 26
s.column_dimensions["B"].width = 8
s.column_dimensions["C"].width = 8

out = "/home/user/ddalgin.github.io/London_Finance_Prospects_6Segments.xlsx"
wb.save(out)
print("Saved", out, "rows:", len(data))
for seg in ORDER:
    print(f"  {sc[seg]:3d}  {seg}")
