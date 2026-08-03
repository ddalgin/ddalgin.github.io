#!/usr/bin/env python3
"""Build London finance conference prospecting list in the target CRM import format."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# (Company, [ (Full Name, Title, Email), ... ])
DATA = [
    ("HSBC", [
        ("Hazem El Taha", "SVP, Group Marketing Technology, International Wealth Premier Banking", "hazem.eltaha@hsbc.co.uk"),
        ("Laura Morgante", "Customer Journey & Proposition Lead", "laura.morgante@hsbc.com"),
        ("Andreas Fenneker", "Associate Director - B2B Marketing", "andreas.fenneker@hsbc.de"),
        ("Steve Thomas", "Head of Marketing Strategy & Global Delivery", "steve.thomas@hsbc.com"),
        ("Nicole German", "Global CMO & Chief Digital Experience Officer", "nicole.german@hsbc.com"),
        ("Alex Leay", "Global Head of Financing Marketing", "alex.leay@hsbc.com"),
        ("William Barraclough", "Head of Product Design, International Wealth & Premier Banking", "william.barraclough@hsbc.com"),
    ]),
    ("Barclays", [
        ("Ben Harding", "Head of Marketing & Engagement Strategy", "ben.harding@barclays.com"),
        ("Dolika Dhar", "Vice President - Senior Compliance Manager", "dolika.dhar@barclays.com"),
        ("Fred Doyhamboure", "Markets - Compliance Surveillance", "Fred.doyhamboure@barclays.com"),
        ("Amit Singh", "VP, Global Product Manager, Integrated Reporting (Digital)", "amit.singh@barclays.com"),
        ("Magnus Deuling", "Global Head of Product Design (MarketsOne - SDP)", "magnus.deuling@barclays.com"),
        ("Lauren Stewart", "VP, Senior Digital Product Owner, Mobile App", "lauren.stewart@barclays.com"),
        ("Priya Mallinson", "Head of Product - Onboarding - Business Banking", "priya.mallinson@barclays.com"),
    ]),
    ("NatWest Group", [
        ("Eleanor Whittaker", "Digital Product Manager", "eleanor.whittaker@natwest.com"),
        ("Maria Lagutina", "Director, FinTech Product Manager - Conversational AI, Chief AI Research Office", "maria.lagutina@natwest.com"),
        ("Wendy Redshaw", "Chief Digital Information Officer, Retail", "wendy.redshaw@natwest.com"),
        ("Robyn Johnston", "Head of CIB Marketing", "robyn.johnston@natwest.com"),
        ("Marty Carroll", "Director, Digital Marketing & Media", "marty.carroll@natwest.com"),
        ("Dasha Lukiniha", "Strategy and Intelligence Lead, Chief Operating Office (Commercial & Institutional)", "Dasha.lukiniha@natwest.com"),
        ("Nicky Mackrell", "Director, Brand Marketing", "nicky.mackrell@natwest.com"),
        ("Matthew Harwood", "Head of Marketing & Engagement Platforms", "matthew.harwood@natwest.com"),
        ("Kate Litler", "Head of Customer Experience & Business Management, Retail Bank", "kate.litler@natwest.com"),
        ("Grant Thomas", "Experience Lead, Branch & Local Banking", "grant.thomas@natwest.com"),
        ("Tanvi Gokhali", "Managing Director, Customer Experience & Engagement", "tanvi.gokhale@natwest.com"),
        ("Amanda Scott", "Head of Digital Contact Channels - Commercial and Institutional Banking", "amanda.scott@natwest.com"),
    ]),
    ("Lloyds Banking Group", [
        ("Aimee Hood", "Conversational Banking - Customer Experience Performance & Strategy", "aimee-munro.hood@lloydsbanking.com"),
        ("Laurie Hodge", "Marketing Strategy Lead", "laurie.hodge@lloydsbanking.com"),
        ("Jay Safdar", "Performance Engine Lead, Consumer Marketing", "jay.sadfar@lloydsbanking.com"),
        ("Ben Logan", "Head of Experience Design - Customer Comms", "ben.logan@lloydsbanking.com"),
        ("Seena Samani", "Head of Marketing Effectiveness", "seena.samani@lloydsbanking.com"),
        ("Ella Davis", "Marketing Campaign Manager - Consumer Banking, Deepening & Retention", "ella.davis@lloydsbanking.com"),
        ("Sangeetha Narasimhan", "Head of GTM Marketing, Merchant Services", "sangeetha.narasimhan1@lloydsbanking.com"),
    ]),
    ("Santander", [
        ("Idoia Muguruza", "Marketing Group VP | Growth & UX", "idoia.muguruza@gruposantander.com"),
        ("Antonio Herranz", "Global AI Transformation, Director", "antonio.herranz@gruposantander.es"),
        ("Bernabe Mohedano Cuadrado", "Head of Brand and Corporate Marketing", "bernabe.mohedano@gruposantander.es"),
        ("Ana Donaire", "Head of Marketing communications & Customer Engagement", "Ana.Donaire@gruposantander.es"),
        ("Silvia Merino Prieto", "SVP, Marketing & Comms Director", "silviamerino@gruposantander.com"),
        ("Andrea Moran Arnedo", "Global Marketing Strategy & Operations | Senior Manager", "andmoran@gruposantander.com"),
        ("Jorge Varela", "Global Digital Channels Planning & Control", "jorge.varela@gruposantander.com"),
        ("David Meza Ayala", "Global AI Product Lead | Agentic & Conversational CEx", "david.meza@gruposantander.com"),
        ("Adam Beardmore", "VP, Head of Product Operations, Digital Channels", "adam.beardmore@gruposantander.com"),
    ]),
    ("BBVA", [
        ("Luis Calleja", "Global Head of Product & Marketing for Digital Banks", "luis.calleja3@bbva.com"),
        ("Miguel Alcala", "BBVA Spark - Managing Director, Head of Europe (Growth Lending & Banking)", "miguel.alcala@bbva.com"),
        ("Horacio Ballestrin", "Head Product & Channel Development @BBVA Spark", "horacio.ballestrin@bbva.com"),
        ("David Arconada", "Global Director of Customer Experience", "david.arconada@bbva.com"),
        ("Elsa Garcia", "Global Marketing Director", "elsa.moya@grupobbva.com"),
        ("Ignacio Gomez", "Head of Marketing and Communications, CIB", "ignacio.gomez@bbva.com"),
        ("Sofia Rodriguez-Sahagun", "Global Discipline Head of Marketing, Digital Sales, Design", "sofia.rodriguez-sahagun@bbva.com"),
        ("Oscar Gomez", "Global Head of Marketing", "oscar.moya@bbva.com"),
        ("Ramon Calderon", "Head of Digital Sales & Marketing Plan at BBVA Spain", "ramon.calderon@bbva.com"),
        ("Walter Rizzi", "Head of Digital Banking, Italy", "walter.rizzi@bbva.com"),
        ("Alfonso Iglesias", "Global Head Digital Sales & Marketing", "alfonso.fernandez.iglesias@bbva.com"),
        ("Martin Eizaga", "Growth, Engagement & BI | BBVA Digital Banking Italy", "Martin.Eizaga@bbva.com"),
        ("Nicola Cecchetto", "Head of Legal Digital Banking Italy", "nicola.cecchetto@bbva.com"),
    ]),
    ("BNP Paribas", [
        ("Damien Cogez", "Senior Lead Product Designer & Design system designer", "cogez.damien@bnpparibas.com"),
        ("Tutku Salihoglu", "Head of Marketing - Soft Mobility", "tutku.salihoglu@bnpparibas.com"),
        ("Mitra Moinfar", "Head of Product - Mobile App Corporate Cards", "mitra.moinfar@bnpparibas.com"),
        ("Romain Rossetti", "Senior Marketing Manager France", "romain.rossetti@bnpparibas-am.com"),
        ("Lisa Bush", "Head of Marketing - Northern Europe", "lisa.bush@bnpparibas-am.com"),
        ("Atman Bentahar", "Head of Omnichannel and Digital | PACE", "Atman.bentahar@bnpparibas.com"),
        ("Soufiane Bachiri", "Organizational Transformation & Customer Experience Expert", "bachiri.soufiane@bnpparibas.com"),
        ("Charlotte Cren", "Head of Customer Journey Team - PACE", "charlotte.cren@bnpparibas.com"),
        ("Deborrah Flament", "Head of Customer Experience and Acquisition / International Retail Banking", "deborrah.flament@bnpparibas.com"),
    ]),
    ("Societe Generale", [
        ("Onsi Kahlaoui", "Head of UX and Product Design", "onsi.kahlaoui@sgcib.com"),
        ("Giorgio Bellati", "Digital Project Director | Chief Product Owner", "giorgio.bellati@socgen.com"),
        ("Pierre-Emmanuel Fraisse", "Chief Product Officer & IT Director (CTO/CIO scope)", "Pierre-Emmanuel.fraisse@socgen.com"),
        ("Tiphaine Massard", "Head of Marketing & Communication - Scalexpert Ecommerce", "tiphaine.massard-godard@socgen.com"),
        ("Salma Radi", "Digital Marketing & Communication Coordinator", "Salma.radi@socgen.com"),
        ("Laetitia Bianchi", "Managing Director - Global Head of Marketing, Global Markets", "laetitia.bianchi@sgcib.com"),
    ]),
    ("Credit Agricole", [
        ("Stephanie Vicat", "Leader Marketing Strategique", "stephanie.vicat@credit-agricole-sa.fr"),
        ("Shivani Bhagat", "Chef de Projet Relationship Marketing", "shivanisunil.bhagat@ca-cib.com"),
        ("Julien Provost-Langot", "Chief Digital Marketing & Customer Experience Officer", "jprovost@ca-cf.fr"),
        ("Marie Gourche", "Chargee de Marketing et Communication Digitale", "marie.schouvey@ca-cib.com"),
        ("Nicolas Renaut", "Head of Marketing & Business Development for Corporate", "nicolas.renaut@credit-agricole-sa.fr"),
        ("Vincent Giraudel", "Digital Marketing and CRM International Project Manager", "vincent.giraudel@credit-agricole-sa.fr"),
        ("Yani Arkat", "Senior AI Product Manager (Agentic & GenAI)", "y.arkat@ca-cib.com"),
        ("Julian Nastase", "COO & Business Manager", "julian.nastase@ca-cib.com"),
        ("James Hanscombe", "Director - Global Markets Advisory Compliance", "james.hanscombe@ca-cib.com"),
    ]),
    ("ING Group", [
        ("Adis Tutundzic", "Global Head of Loyalty & Merchant Services", "adis.tutundzic@ing.com"),
        ("Michel Drupsteen", "Transformation lead GenAI", "michel.drupsteen@ing.com"),
        ("Yvon Martin", "Global Lead, Brand Strategy & Marketing", "yvon.martin@ing.com"),
        ("Arjen Hoekstra", "Global Lead for Digital Marketing & Sales, ING Business Banking", "arjen.hoekstra@ing.com"),
        ("Willem Speelman", "Head of Marketing Digital Transformation & Customer", "willem.speelman@ing.com"),
        ("Amit Vats", "Strategy & Transformation - Retail Business Banking", "amit.vats@ing.com"),
        ("Robert Rugebregt", "Customer Journey Expert | Product Owner | Product Manager", "robert.rugebregt@ing.com"),
        ("Omer Cerrahoglu", "Head of Customer Insights and Connectivity - Payments", "omer.cerrahoglu@ing.com"),
        ("Jaap Burggraaff", "Chief CX - Customer Experience Expert", "jaap.burggraaff@ing.com"),
        ("Niels Stroeve", "Customer Journey Expert - Digital & Customer Experience", "niels.stroeve@ing.com"),
        ("Aksidan Kukalev", "Director - Regulatory Operations", "aksidan.kukalev@ing.com"),
        ("Murat Ayrancioglu", "Payment Cards Global Vendor Manager", "murat.ayrancioglu@ing.com"),
        ("Arjen Noordeman", "Global Senior Brand Identity & Design Expert", "arjen.noordeman@ing.com"),
    ]),
    ("ABN AMRO", [
        ("Marijke Goudberg", "Procurement Category Manager Marketing, Sponsoring", "marijke.goudberg@nl.abnamro.com"),
        ("Benko Vermolen", "Head of Operations Delivery & Expertise Center", "benko.vermolen@nl.abnamro.com"),
        ("Yorick Naeff", "Head of Innovation", "yorick.naeff@nl.abnamro.com"),
        ("Lisa Dubach", "Head of Product: Human Channels", "lisa.dubach@nl.abnamro.com"),
        ("Erik Wijnen", "Head of Credits Operations & Debt Collections", "erik.wijnen@nl.abnamro.com"),
        ("Aude Josset", "Director, Marketing & Communications", "aude.josset@fr.abnamro.com"),
    ]),
    ("Millennium BCP", [
        ("Joao Trigueiros", "Head of Customer Strategy & Partnerships - Executive", "joao.forjaztrigueiros@millenniumbcp.pt"),
        ("Andre Correia", "Head of Digital Marketing and Communication", "andrecorreia@millenniumbcp.pt"),
        ("Rachel Ferreira", "Senior Manager - Data & AI Transformation | Reporting", "rachel.ferreira@millenniumbcp.pt"),
        ("Isabel Santos", "Head of Business Support Division at Corporate Market", "isantos@millenniumbcp.pt"),
        ("Filipe Marques", "Head of Payment Products Marketing", "fmarques@millenniumbcp.pt"),
        ("Fernanda Aguiar", "Unit Director @Private Banking Marketing Department", "fernanda.aguiar@millenniumbcp.pt"),
        ("Jose Nunes", "Digital Transformation Director - Consumer Lending", "jose.nunes@millenniumbcp.pt"),
        ("Pedro Mendes", "Diretor Experiencia do Cliente/Modelos de Satisfacao", "pedromendes@millenniumbcp.pt"),
        ("Miguel Garcia Almeida", "Digital Content Lead | UX & Product Content | Copywriting & Conversion", "miguel.almeida@millenniumbcp.pt"),
        ("Pedro Beija", "Head Of Operations", "pedro.beija@millenniumbcp.pt"),
        ("Joao Jesus", "Head of Data & AI Transformation Office", "joao.jesus@millenniumbcp.pt"),
        ("Jose Farinha", "B2B Integration Lead | Banking & Treasury Automation", "jose.farinha@millenniumbcp.pt"),
    ]),
    ("Deutsche Bank", [
        ("Snehal Pandya", "Director - Data & AI Change Execution Leader", "sam.pandya@db.com"),
        ("Stephen Bell", "Global Head of Trade Finance and Lending Operations", "stephen.bell@db.com"),
        ("Marcel Everts", "Transformation: Business Strategy & Process Excellence", "marcel.everts@db.com"),
        ("Raquel Carrillo", "Head Of Marketing", "raquel.carrillo@db.com"),
        ("Nicki Saini", "Associate Vice President, Event Marketing", "Nicki.saini@db.com"),
        ("Brigitte Koetting", "Chief Marketing Owner API Banking", "brigitte.koetting@db.com"),
        ("Merle Meier-Holsten", "Managing Director Marketing Personal Banking Private Bank", "merle.meier-holsten@db.com"),
        ("Priyanka Dayal", "Global Marketing Specialist", "priyanka.dayal@db.com"),
        ("Kim Kranz", "VP, Global Head of Content Marketing", "kim-maria.kranz@db.com"),
        ("Holger Fischer", "Global Head of Brand Strategy", "holger.fischer@db.com"),
        ("Mina Saidze", "Vice President Tech Product Management - Digital Sales", "mina.saidze@db.com"),
        ("David Dommel", "Director I Head of Performance Marketing", "david.dommel@db.com"),
        ("Lisa Barahona", "Lead Technical Product Manager", ""),
        ("Morgane Constanty", "Head of AI Product Management, Corporate Bank Chief", "Morgane.constanty@db.com"),
        ("Marcin Belinski", "VP Product", "marcin.belinski@db.com"),
    ]),
    ("CaixaBank", [
        ("Carles Aznar Carrique", "Product Lead (Bank Loans and Overdrafts)", "caznar@caixabankpc.com"),
        ("German Mazza", "Digital Transformation Manager", "gdmontero@caixabank.com"),
        ("Sergio Ruiz", "Product Design Lead", "scedeno@caixabank.com"),
        ("Angels Valls Benitez", "Digital Communication Transformation Director", "avalls@caixabank.com"),
        ("Eduardo Mallart", "Marketing Business Manager", "emallart@caixabank.com"),
        ("Joan Gaitx", "SVP, CIB solutions", "jgaitx@caixabank.com"),
        ("Genis Barnich Fonolla", "Adtech Lead - Marketing Technology", "gbarnich@caixabank.com"),
        ("Sergio Ponce", "Director Marketing Client Experience", "sergio.gutierrez.p@caixabank.com"),
        ("Ruben Morin Martin", "Head of Tech Digital Marketing & Digital Platforms", "rmorinm@caixabanktech.com"),
        ("Jordi Guaus", "Head of Digital Marketing", "jguaus@caixabank.com"),
        ("Carla Manonelles", "Director of Customer Experience & Communication", "cmanonelles@caixabankpc.com"),
        ("Ignacio Molinero", "Director Customer Experience", "imolineroh@caixabank.com"),
    ]),
    ("Revolut", [
        ("Vladimir Kovtun", "Product Lead", "vladimir.kovtun@revolut.com"),
        ("Madeline White", "Senior Product Marketing Manager", "m.white@revolut.com"),
        ("Eleanor Taylor", "Senior Creative Marketing Manager", "eleanor.taylor@revolut.com"),
        ("Kuba Fast", "CEO, Revolut Bank EU", "kuba.fast@revolut.com"),
        ("Leila Sakhnini", "Performance Marketing Manager", "leila.sakhnini@revolut.com"),
        ("Sofya Tretyakova", "Marketing Director", "sofya.tretyakova@revolut.com"),
        ("Sonia Calderon", "Mentor in digital marketing", "sonia.calderon@revolut.com"),
        ("Theau Guillot", "Growth Marketing Manager", "theau.guillot@revolut.com"),
        ("Marco Tracogna", "Head of Performance Marketing Europe", "marco.tracogna@revolut.com"),
        ("Sergei Puzin", "Senior Product Owner (Technical)", "sergey.puzin@revolut.com"),
        ("Irina Zhukova", "Director Executive (Marketing)", "irina.zhukova@revolut.com"),
        ("Edouard Daunizeau", "Lead Product Marketing Manager", "Edouard.daunizeau@revolut.com"),
        ("Wes Brannigan", "Head of Operations, UK", "wes.brannigan@revolut.com"),
        ("Cedric Sevin", "Head of CX Product Operations", "Cedric.sevin@revolut.com"),
        ("Afonso Coelho", "COO Office - Lead Operations Manager", "a.amorim@revolut.com"),
        ("Pietro Dalpane", "Head of Global Operations - Revolut Savings", "pietro.dalpane@revolut.com"),
    ]),
    ("Wise", [
        ("Ziggy Pragier", "Global Head of Screening Operations", "ziggy.pragier@wise.com"),
        ("Nilan Peiris", "Chief Product Officer", "nilan.peiris@wise.com"),
        ("Josh Payton", "VP, Design", "josh.payton@wise.com"),
        ("Sonia-Carla Ciocirlie", "Director of Business Operations", "sonia.cialicu@wise.com"),
        ("Samuele Erbi", "Head of Third Party Management & Procurement", "samuele.erbi@wise.com"),
        ("Magda Panfil", "Senior Director Product Operations", "magda.panfil@wise.com"),
        ("Ruth Chadwick", "Global Director of Brand, Creative and Strategy", "ruth.chadwick@wise.com"),
        ("Chetna Shastry", "Head of Operations - Global Product", "chetna.shastry@wise.com"),
        ("Mike Shebalkov", "Product Director", "mike.shebalkov@wise.com"),
        ("Ayesha Kadan", "Head of Merchant Growth & Ad Operations", "ayesha.kadan@wise.com"),
        ("Samara Thomas", "Head of Creative Operations", "samara.thomas@wise.com"),
        ("Cian Weeresinghe", "Chief Marketing Officer", "cian.weeresinghe@wise.com"),
        ("Alessandro Battaglia", "Head of Paid Acquisition", "alessandro.battaglia@wise.com"),
        ("Iona Carter", "VP Brand & Marketing Strategy", "Iona.carter@wise.com"),
    ]),
    ("Monzo", [
        ("Natalie Malevsky", "Head of Brand Strategy", "Natalie.malevsky@monzo.com"),
        ("Nicole Christensen", "EU Expansion Marketing Director", "nicolechristensen@monzo.com"),
        ("Chris Mucklow-Norell", "Head of Brand Marketing", "chrisnorell@monzo.com"),
        ("Emily Suter", "Director of Brand Marketing", "emilysuter@monzo.com"),
        ("Ben Lawrence", "Director of Product Marketing", "benlawrence@monzo.com"),
        ("Rannie Powell", "Director of Marketing Tech, AI & Insights", "ranniepowell@monzo.com"),
        ("Katie Brown", "Head of Business and International Hiring", "katiegormley@monzo.com"),
        ("Karen Addy", "Standards & Governance Director - Customer Operations", "karenaddy@monzo.com"),
        ("Arundhathi Desai", "Commercial & Operations Director, Core Banking", "arundhathidesai@monzo.com"),
        ("Rachad Khoury", "CX & Operations", "Rachad.khoury@monzo.com"),
        ("Jordan Shwide", "VP, General Manager, Business Banking", "jordanshwide@monzo.com"),
        ("Katrina O'Donnell", "Director Customer Engagement & CRM", "katodonnell@monzo.com"),
        ("Nicole Rudland", "Director Customer Operations", "Nicole.rudland@monzo.com"),
        ("Rich Harris", "Director of Vulnerable Customers", "richharris@monzo.com"),
        ("Shelley Malton", "Group Chief Operating Officer", "shelleymalton@monzo.com"),
    ]),
    ("Starling Bank", [
        ("Jessie Broome", "Head of AML, Screening Operations & Strategy", "jessie.harringtonbroome@starlingbank.com"),
        ("Joe Bakowski", "Group Head of Procurement & Vendor Management", "joe.bakowski@starlingbank.com"),
        ("Chris Ross", "Head of Technology Operations", "chris.ross@starlingbank.com"),
        ("Emma Stanning", "Chief of Staff to CMO & Director of Marketing Operations", "emma.stanning@starlingbank.com"),
        ("Daniel Cosentino", "Head of Projects", "daniel.cosentino@starlingbank.com"),
        ("Lisa Cunico", "Head of Production", "lisa.cunico@starlingbank.com"),
        ("Michele Rousseau", "Chief Marketing Officer", "michele.rousseau@starlingbank.com"),
        ("Raghu Narula", "Chief Customer & Banking Officer", "raghu.narula@starlingbank.com"),
        ("Indiana Matine", "Director of Brand Strategy", "indiana.matine@starlingbank.com"),
        ("Luke Farrell", "Customer and Banking, Head of Customer Journey Management", "luke.farrell@starlingbank.com"),
        ("Evangelos Rakovitis", "Head of Retail Banking Solutions", "evangelos.rakovitis@starlingbank.com"),
        ("Sami Kade", "Banking Product Director", "sami.kade@starlingbank.com"),
        ("Aaron Shaw", "Director of Product & Engineering", "aaron.shaw@starlingbank.com"),
    ]),
    ("N26", [
        ("Luis Franqueira", "Head of Engineering, Developer Experience", "luis.franqueira@n26.com"),
        ("Prithviraj Sarkar", "Head of Product - Customer Identity", "prithvi.sarkar@n26.com"),
        ("Catalina Garcia", "Head of Marketing & Communication Iberia, DACH", "catalina.garcia@n26.com"),
        ("Covadonga Calvo", "Director Marketing Management", "covadonga.garcia@n26.com"),
        ("Laura Cristaldi", "Head of Business Domain | Customer Service & Ops (AI)", "laura.cristaldi@n26.com"),
        ("Niklas Oertel", "Head of Growth - Germany", "niklas.oertel@n26.com"),
        ("Pierre Plessis", "Creative Director - Copy, Editorial Content, and Localization", ""),
        ("Will Sorby", "Director of Product: Investments, Lending & Savings", "will.sorber@n26.com"),
        ("Katharina Pollak", "Director of Lifecycle (CRM)", "katharina.pollak@n26.com"),
        ("Crystal Goh", "Global Head Of Communications", "crystal.goh@n26.com"),
        ("Holger Boehmer", "Managing Director N26 Operations GmbH", "holger.boehmer@n26.com"),
        ("Siddharth Paul", "Head of Banking Operations", "siddharth.paul@n26.com"),
        ("Christopher Poots", "Executive Creative Director", "christopher.poots@n26.com"),
    ]),
    ("Trade Republic", [
        ("Marcus Letsch", "Head of Growth Marketing", "marcus.letsch@traderepublic.com"),
        ("Nina Jetter-Schmidt", "Director Corporate Development", "nina.jetter@traderepublic.com"),
        ("Alexander Blum", "Head of AFC Compliance", "alexander.blum@traderepublic.com"),
        ("Julian Collin", "VP Growth & General Manager International Markets", "julian.collin@traderepublic.com"),
        ("Ulrich Schroeer", "Director Risk & Compliance", "ulrich.schroeer@traderepublic.com"),
        ("Daniel Tordable", "Affiliate & Influencer Marketing Manager", "daniel.montero@traderepublic.com"),
        ("Kirill Novikov", "Senior Product Manager", "Kirill.novikov@traderepublic.com"),
        ("Ghazal Azari", "Product Operations", "ghazal.azari@traderepublic.com"),
        ("Dora Trostanetsky", "Growth Marketing", "Dora.trostanetsky@traderepublic.com"),
        ("Felix Jamestin", "Product", "felix.jamestin@traderepublic.com"),
        ("Laura Eggers", "Global People Operations Lead", "laura.eggers@traderepublic.com"),
        ("Victor Kovatsenko", "AI/ML Product Manager (AML & Fraud)", "victor.kovatsenko@traderepublic.com"),
    ]),
]

# Company -> HQ country (used only to pre-fill the Country column as a helpful default)
COMPANY_COUNTRY = {
    "HSBC": "United Kingdom", "Barclays": "United Kingdom", "NatWest Group": "United Kingdom",
    "Lloyds Banking Group": "United Kingdom", "Revolut": "United Kingdom", "Wise": "United Kingdom",
    "Monzo": "United Kingdom", "Starling Bank": "United Kingdom",
    "Santander": "Spain", "BBVA": "Spain", "CaixaBank": "Spain",
    "BNP Paribas": "France", "Societe Generale": "France", "Credit Agricole": "France",
    "ING Group": "Netherlands", "ABN AMRO": "Netherlands",
    "Millennium BCP": "Portugal",
    "Deutsche Bank": "Germany", "N26": "Germany", "Trade Republic": "Germany",
}

from country_overrides import OVERRIDES

HEADERS = [
    "First Name", "Last Name", "Job Title", "Company", "Phone", "Email", "Country",
    "State (US)", "Source", "Lead Status", "Privacy Policy Accepted", "Score",
    "Owner email address", "Lead Queue", "Main Marketing Language", "Lead Type 1",
    "Lead Sub Type", "Vertical", "Brand Preference",
    # Helper columns (NOT part of the A-S CRM import - remove before importing if needed)
    "Country Confidence", "Country Evidence",
]

# Multi-word first names that should be kept together
COMPOUND_FIRST = {"Pierre-Emmanuel", "Sonia-Carla", "Kim", "Chris", "Nina"}


def split_name(full):
    parts = full.split()
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


wb = Workbook()
ws = wb.active
ws.title = "Prospects"

FONT = "Arial"
header_font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor="1F4E78")
cell_font = Font(name=FONT, size=10)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header row
for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=1 + 0)  # placeholder, set below
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

verified_fill = PatternFill("solid", fgColor="E2EFDA")   # light green
assumed_fill = PatternFill("solid", fgColor="FCE4D6")    # light orange

r = 2
missing_keys = []
n_verified = 0
n_assumed = 0
for company, people in DATA:
    hq_country = COMPANY_COUNTRY.get(company, "")
    for full, title, email in people:
        fn, ln = split_name(full)
        key = f"{company}||{full}"
        ov = OVERRIDES.get(key)
        if ov is None:
            missing_keys.append(key)
            country, confidence, evidence = hq_country, "Assumed", "no override found (HQ default)"
        else:
            country, confidence, evidence = ov
        if confidence == "Verified":
            n_verified += 1
        else:
            n_assumed += 1
        row_vals = [fn, ln, title, company, "", email, country, "", "", "", "", "",
                    "", "", "", "", "", "Financial Services", "", confidence, evidence]
        for c, v in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        # shade the Country + Confidence cells by confidence level
        conf_fill = verified_fill if confidence == "Verified" else assumed_fill
        ws.cell(row=r, column=7).fill = conf_fill    # Country
        ws.cell(row=r, column=20).fill = conf_fill   # Country Confidence
        r += 1

if missing_keys:
    print("WARNING: missing override keys:")
    for k in missing_keys:
        print("   ", k)
print(f"Country confidence -> Verified: {n_verified}, Assumed: {n_assumed}")

# Freeze header + autofilter
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:U{r-1}"

# Column widths
widths = {
    "A": 16, "B": 20, "C": 52, "D": 22, "E": 14, "F": 40, "G": 16, "H": 12,
    "I": 16, "J": 14, "K": 20, "L": 8, "M": 22, "N": 14, "O": 20, "P": 14,
    "Q": 16, "R": 20, "S": 18, "T": 18, "U": 60,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 30

# ---- Notes sheet -------------------------------------------------
nt = wb.create_sheet("Read me")
notes = [
    ("London Finance Conference - Prospecting List", True),
    ("", False),
    (f"Total contacts: {r-2}   |   Companies: {len(DATA)}", False),
    ("Format: matches your CRM import template (columns A-S).", False),
    ("", False),
    ("Columns populated from the source list:", True),
    ("  - First Name, Last Name  (split from the full name)", False),
    ("  - Job Title", False),
    ("  - Company", False),
    ("  - Email", False),
    ("  - Country  (LinkedIn/web cross-reference - see Country Confidence)", False),
    ("  - Vertical  (set to 'Financial Services' for all rows)", False),
    ("", False),
    ("Country column - how it was determined:", True),
    ("  Each contact's name + title + company was searched on the web to find their", False),
    ("  matching LinkedIn profile, then the profile's location was used for Country.", False),
    ("  Two helper columns (T and U, OUTSIDE the A-S import range) record this:", False),
    ("    - Country Confidence:  'Verified' (green) = a matching profile showed a real", False),
    ("      location signal (explicit city/country or a country-coded LinkedIn subdomain);", False),
    ("      'Assumed' (orange) = no confident match, fell back to the company HQ country.", False),
    ("    - Country Evidence:  the specific signal used for each row.", False),
    ("  NOTE: remove columns T and U before importing if your CRM expects only A-S.", False),
    ("  All countries fall within your 6: UK, France, Germany, Netherlands, Spain, Portugal", False),
    ("  EXCEPTIONS found outside those 6 (please review - see orange/notes):", False),
    ("    - Walter Rizzi & Nicola Cecchetto (BBVA) -> Italy (Milan)", False),
    ("    - Martin Eizaga (BBVA) -> Italy (likely; BBVA Digital Banking Italy)", False),
    ("    - Kuba Fast (Revolut) -> CEO of Revolut's Lithuania entity (Vilnius); ex-Chase UK,", False),
    ("      physical base ambiguous UK/Lithuania - marked UK/Assumed, NEEDS REVIEW.", False),
    ("  ~21 contacts (rest of N26 + all Trade Republic) could not be searched (hit the", False),
    ("  session's 200-search limit); they use the Berlin/Germany HQ default (Assumed).", False),
    ("", False),
    ("Columns intentionally left blank for the marketing team to complete", True),
    ("(these depend on your CRM picklists / campaign setup):", False),
    ("  Phone, State (US), Source, Lead Status, Privacy Policy Accepted, Score,", False),
    ("  Owner email address, Lead Queue, Main Marketing Language, Lead Type 1,", False),
    ("  Lead Sub Type, Brand Preference", False),
    ("", False),
    ("Notes on the data:", True),
    ("  - Two contacts had no email in the source (Lisa Barahona / Deutsche Bank,", False),
    ("    Pierre Plessis / N26); their Email cells are blank.", False),
    ("  - One source email had a typo (kuba,fast@revolut.com) - corrected to", False),
    ("    kuba.fast@revolut.com. Please confirm before sending.", False),
    ("  - Accented characters in some French/Portuguese titles were simplified to", False),
    ("    plain ASCII for clean CRM import.", False),
]
title_font = Font(name=FONT, bold=True, size=13)
bold_font = Font(name=FONT, bold=True, size=10)
reg_font = Font(name=FONT, size=10)
for i, (text, is_bold) in enumerate(notes, start=1):
    cell = nt.cell(row=i, column=1, value=text)
    if i == 1:
        cell.font = title_font
    elif is_bold:
        cell.font = bold_font
    else:
        cell.font = reg_font
nt.column_dimensions["A"].width = 80

out = "/home/user/ddalgin.github.io/London_Finance_Prospect_List.xlsx"
wb.save(out)
print(f"Saved {out} with {r-2} contacts across {len(DATA)} companies")
